from flask import Flask, jsonify, request
from openpyxl import load_workbook
from openpyxl.styles import PatternFill
from datetime import datetime
from dateutil.relativedelta import relativedelta
import logging
from datetime import timedelta
import os

app = Flask(__name__)
file_path = 'tasks.xlsx'

# Set up logging
logging.basicConfig(level=logging.DEBUG)

def read_tasks():
    try:
        if not os.path.exists(file_path):
            logging.error(f"File {file_path} does not exist in directory {os.getcwd()}")
            return {"all": [], "completed": []}

        logging.debug(f"Loading workbook from {file_path}")
        wb = load_workbook(file_path)
        
        logging.debug(f"Available sheets: {wb.sheetnames}")
        
        if 'AllTasks' not in wb.sheetnames or 'CompletedTasks' not in wb.sheetnames:
            logging.error(f"Required sheets 'AllTasks' or 'CompletedTasks' not found in {file_path}")
            return {"all": [], "completed": []}

        all_tasks_sheet = wb['AllTasks']
        completed_tasks_sheet = wb['CompletedTasks']

        tasks = {
            "all": [],
            "completed": []
        }

        def parse_tasks(sheet, include_comments=False, include_status=False):
            task_list = []
            row_count = sum(1 for _ in sheet.iter_rows(min_row=2))
            logging.debug(f"Parsing {sheet.title} with {row_count} data rows")
            for row_idx, row in enumerate(sheet.iter_rows(min_row=2, values_only=True), start=2):
                try:
                    if row[0] is None or row[1] is None or row[2] is None:
                        logging.warning(f"Skipping invalid row {row_idx} in {sheet.title}: {row}")
                        continue
                    task = {
                        "name": str(row[0]),
                        "priority": str(row[1]),
                        "date_added": str(row[2])
                    }
                    if include_comments and len(row) > 3:
                        task["comment"] = str(row[3] or "")
                    if include_status and len(row) > 4:
                        task["status"] = str(row[4] or "")
                    task_list.append(task)
                except Exception as e:
                    logging.warning(f"Error parsing row {row_idx} in {sheet.title}: {row}, error: {e}")
            return task_list

        tasks["all"] = sorted(
            parse_tasks(all_tasks_sheet),
            key=lambda x: (x['priority'], x['date_added'])
        )
        tasks["completed"] = parse_tasks(completed_tasks_sheet, include_comments=True, include_status=True)

        logging.debug(f"Loaded {len(tasks['all'])} tasks from AllTasks and {len(tasks['completed'])} from CompletedTasks")
        return tasks
    except Exception as e:
        logging.error(f"Error reading tasks from {file_path}: {e}")
        return {"all": [], "completed": []}

def mark_task_completed(task_name, comment="", status="Completed"):
    try:
        wb = load_workbook(file_path)
        all_tasks_sheet = wb['AllTasks']
        completed_tasks_sheet = wb['CompletedTasks']

        for row in all_tasks_sheet.iter_rows(min_row=2, values_only=False):
            if row[0].value == task_name:
                completed_row = [cell.value for cell in row]
                completed_row[2] = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
                completed_row.append(comment)
                completed_row.append(status)
                completed_tasks_sheet.append(completed_row)
                
                if status == "Cannot Complete":
                    new_row = completed_tasks_sheet.max_row
                    for cell in completed_tasks_sheet[new_row]:
                        cell.fill = PatternFill(start_color="FF9999", end_color="FF9999", fill_type="solid")
                
                all_tasks_sheet.delete_rows(row[0].row)
                break

        wb.save(file_path)
    except Exception as e:
        logging.error(f"Error marking task completed: {e}")

def get_recent_pass_fail():
    try:
        wb = load_workbook(file_path)
        pass_fail_sheet = wb['Pass Fail']

        most_recent_row = None
        for row in pass_fail_sheet.iter_rows(min_row=2, values_only=True):
            if most_recent_row is None or row[0] > most_recent_row[0]:
                most_recent_row = row

        if most_recent_row:
            return {
                "date": most_recent_row[0],
                "pass": most_recent_row[1],
                "fail": most_recent_row[2]
            }
        return {}
    except Exception as e:
        logging.error(f"Error getting pass/fail data: {e}")
        return {}

def get_turn_around_time():
    try:
        wb = load_workbook(file_path)
        if 'Turn Around Time' not in wb.sheetnames:
            logging.error("Turn Around Time sheet not found")
            return []

        turn_around_sheet = wb['Turn Around Time']
        five_months_ago = datetime.now() - relativedelta(months=5)
        
        data = []
        for row in turn_around_sheet.iter_rows(min_row=2, values_only=True):
            try:
                if isinstance(row[0], datetime):
                    row_date = row[0]
                else:
                    row_date = datetime.strptime(str(row[0]), "%Y-%m-%d %H:%M:%S")
                
                if row_date >= five_months_ago and isinstance(row[1], (int, float)):
                    data.append({
                        "date": row_date.strftime("%Y-%m-%d %H:%M:%S"),
                        "turn_around_time": float(row[1])
                    })
                    logging.debug(f"Processed row: {row}")
            except (ValueError, TypeError) as e:
                logging.warning(f"Skipping invalid row: {row}, error: {e}")
                continue

        monthly_data = {}
        for item in data:
            date = datetime.strptime(item['date'], "%Y-%m-%d %H:%M:%S")
            month_key = date.strftime("%Y-%m")
            if month_key not in monthly_data:
                monthly_data[month_key] = {"times": [], "date": date.strftime("%Y-%m-01 00:00:00")}
            monthly_data[month_key]["times"].append(item['turn_around_time'])

        result = []
        for month_key, info in monthly_data.items():
            avg_time = sum(info["times"]) / len(info["times"]) if info["times"] else 0.0
            result.append({
                "date": info["date"],
                "turn_around_time": round(avg_time, 2)
            })

        result.sort(key=lambda x: x['date'])
        logging.debug(f"Turn around time data: {result}")
        return result[-5:]
    except Exception as e:
        logging.error(f"Error getting turn around time: {e}")
        return []

def get_open_close_monthly():
    try:
        wb = load_workbook(file_path)
        if 'Open Close Monthly' not in wb.sheetnames:
            logging.error("Open Close Monthly sheet not found")
            return []

        open_close_sheet = wb['Open Close Monthly']
        five_months_ago = datetime.now() - relativedelta(months=5)
        
        data = []
        for row in open_close_sheet.iter_rows(min_row=2, values_only=True):
            try:
                if isinstance(row[0], datetime):
                    row_date = row[0]
                else:
                    row_date = datetime.strptime(str(row[0]), "%Y-%m-%d %H:%M:%S")
                
                if row_date >= five_months_ago and isinstance(row[1], (int, float)) and isinstance(row[2], (int, float)):
                    data.append({
                        "date": row_date.strftime("%Y-%m-%d %H:%M:%S"),
                        "open": int(row[1]),
                        "close": int(row[2])
                    })
                    logging.debug(f"Processed open/close row: {row}")
            except (ValueError, TypeError) as e:
                logging.warning(f"Skipping invalid open/close row: {row}, error: {e}")
                continue

        monthly_data = {}
        for item in data:
            date = datetime.strptime(item['date'], "%Y-%m-%d %H:%M:%S")
            month_key = date.strftime("%Y-%m")
            if month_key not in monthly_data:
                monthly_data[month_key] = {"open": 0, "close": 0, "date": date.strftime("%Y-%m-01 00:00:00")}
            monthly_data[month_key]["open"] += item['open']
            monthly_data[month_key]["close"] += item['close']

        result = []
        for month_key, info in monthly_data.items():
            result.append({
                "date": info["date"],
                "open": info["open"],
                "close": info["close"]
            })

        result.sort(key=lambda x: x['date'])
        logging.debug(f"Open close monthly data: {result}")
        return result[-5:]
    except Exception as e:
        logging.error(f"Error getting open close monthly data: {e}")
        return []

def get_types():
    try:
        wb = load_workbook(file_path)
        if 'Types' not in wb.sheetnames:
            logging.error("Types sheet not found")
            return []

        types_sheet = wb['Types']
        data = []
        for row in types_sheet.iter_rows(min_row=2, values_only=True):
            try:
                if row[0] is not None and isinstance(row[1], (int, float)):
                    data.append({
                        "type": str(row[0]),
                        "qty": int(row[1])
                    })
                    logging.debug(f"Processed types row: {row}")
                else:
                    logging.warning(f"Skipping invalid types row: {row}")
            except (ValueError, TypeError) as e:
                logging.warning(f"Skipping invalid types row: {row}, error: {e}")
                continue

        logging.debug(f"Types data: {data}")
        return data
    except Exception as e:
        logging.error(f"Error getting types data: {e}")
        return []

def get_tasks_completed_this_week():
    try:
        wb = load_workbook(file_path)
        completed_tasks_sheet = wb['CompletedTasks']
        
        today = datetime.now()
        start_of_week = today - timedelta(days=today.weekday())
        end_of_week = start_of_week + timedelta(days=6, hours=23, minutes=59, seconds=59)
        
        count = 0
        for row in completed_tasks_sheet.iter_rows(min_row=2, values_only=True):
            try:
                if isinstance(row[2], datetime):
                    completed_date = row[2]
                else:
                    completed_date = datetime.strptime(str(row[2]), "%Y-%m-%d %H:%M:%S")
                
                if start_of_week <= completed_date <= end_of_week:
                    count += 1
            except (ValueError, TypeError) as e:
                logging.warning(f"Skipping invalid completed task row: {row}, error: {e}")
                continue
        
        logging.debug(f"Tasks completed this week: {count}")
        return {"count": count}
    except Exception as e:
        logging.error(f"Error getting tasks completed this week: {e}")
        return {"count": 0}

@app.route('/tasks', methods=['GET'])
def get_tasks():
    tasks = read_tasks()
    return jsonify(tasks)

@app.route('/tasks/complete', methods=['POST'])
def complete_task():
    data = request.json
    task_name = data.get('name')
    comment = data.get('comment', '')
    status = data.get('status', 'Completed')
    mark_task_completed(task_name, comment, status)
    return '', 204

@app.route('/pass_fail', methods=['GET'])
def pass_fail():
    data = get_recent_pass_fail()
    return jsonify(data)

@app.route('/turn_around_time', methods=['GET'])
def turn_around_time():
    data = get_turn_around_time()
    return jsonify(data)

@app.route('/open_close_monthly', methods=['GET'])
def open_close_monthly():
    data = get_open_close_monthly()
    return jsonify(data)

@app.route('/types', methods=['GET'])
def types():
    data = get_types()
    return jsonify(data)

@app.route('/tasks_completed_this_week', methods=['GET'])
def tasks_completed_this_week():
    data = get_tasks_completed_this_week()
    return jsonify(data)

if __name__ == '__main__':
    app.run(port=5000)
